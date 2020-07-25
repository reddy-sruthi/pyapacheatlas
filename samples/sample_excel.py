import json
import os

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

# PyApacheAtlas packages
from pyapacheatlas.auth import ServicePrincipalAuthentication # Connect to Atlas via a Service Principal
from pyapacheatlas.core import AtlasClient # Communicate with your Atlas server
from pyapacheatlas.scaffolding import column_lineage_scaffold # Create dummy types
from pyapacheatlas.scaffolding.templates import excel_template # Create the excel template file to be populated
from pyapacheatlas.readers import from_excel # Read in the populated excel file.
from pyapacheatlas.readers.excel import ExcelConfiguration # Customize header prefixes (e.g. "Sink" rather than "Target") and sheet names

if __name__ == "__main__":

    # Authenticate against your Atlas server
    oauth = ServicePrincipalAuthentication(
        tenant_id = os.environ.get("TENANT_ID"),
        client_id = os.environ.get("CLIENT_ID"),
        client_secret = os.environ.get("CLIENT_SECRET")
    )
    atlas_client = AtlasClient(
        endpoint_url =  os.environ.get("ENDPOINT_URL"),
        authentication = oauth
    )

    # Create an empty excel template to be populated
    file_path = "./atlas_excel_template.xlsx"
    excel_template(file_path)

    wb = load_workbook(file_path)
    table_sheet = wb.get_sheet_by_name("Tables")
    columns_sheet = wb.get_sheet_by_name("Columns")

    # TABLE Sheet SCHEMA
    # "Target Table", "Target Type", "Target Classifications",
    # "Source Table", "Source Type", "Source Classifications",
    # "Process Name", "Process Type"
    # LIMITATION: Does not support multiple outputs from same process
    tables_to_load = [
        ["DestTable01", "demo_table", None, "SourceTable01", "demo_table", None, "Daily_ETL", "demo_process"],
        ["DestTable01", "demo_table", None, "SourceTable02", "demo_table", None, "Daily_ETL", "demo_process"],
        ["DestTable02", "demo_table", None, "SourceTable03", "demo_table", None, "Weekly_ETL", "demo_process"],
        ["DestTable03", "demo_table", None, None, None, None, "Stored_Proc:Do_Something", "demo_process"]
    ]
    # COLUMNS Sheet SCHEMA
    # "Target Table", "Target Column", "Target Classifications",
    # "Source Table", "Source Column", "Source Classifications",
    # "Transformation"
    columns_to_load = [
        ["DestTable01", "dest_c01", None, "SourceTable01", "source_c01", None, None],
        ["DestTable01", "dest_c02", None, "SourceTable01", "source_c02", None, None],
        # Demonstrate the ability to merge multiple columns
        ["DestTable01", "dest_combo01", None, "SourceTable01", "source_c03", None, "source_c03 + source_c04"],
        ["DestTable01", "dest_combo01", None, "SourceTable02", "source_c04", None, "source_c03 + source_c04"],
        # Demonstrate a simple, straightforward table with classifications
        ["DestTable02", "dest_c03", None, "SourceTable03", "source_c05", "PII", None],
        ["DestTable02", "dest_c04_express", None, None, None, None, "CURRENT_TIMESTAMP()"],
        # Demonstrate a table with no sources at all
        ["DestTable03", "dest_c100_express", None, None, None, None, "CURRENT_TIMESTAMP()"],
        ["DestTable03", "dest_c101_express", None, None, None, None, "RAND(100)"],
        ["DestTable03", "dest_c102_notransform", None, None, None, None, None],
    ]
    
    # Populate the excel template with samples above
    table_row_counter = 0
    for row in table_sheet.iter_rows(min_row=2, max_col=8, max_row= len(tables_to_load)+1):
        for idx, cell in enumerate(row):
            cell.value = tables_to_load[table_row_counter][idx]
        table_row_counter += 1
    
    column_row_counter = 0
    for row in columns_sheet.iter_rows(min_row=2, max_col=7, max_row= len(columns_to_load)+1):
        for idx, cell in enumerate(row):
            cell.value = columns_to_load[column_row_counter][idx]
        column_row_counter += 1
        
    wb.save(file_path)

    # Generate the base atlas type defs for the demo of table and column lineage
    atlas_type_defs = column_lineage_scaffold("demo", useColumnMapping=True)
    # Alternatively, you can get all atlas types via...
    # all_type_defs = client.get_all_typedefs()

    # Upload scaffolded type defs and view the results of upload
    _upload_typedef = client.upload_typedefs(atlas_type_defs)
    print(json.dumps(_upload_typedef,indent=2))

    # Instantiate some required objects and generate the atlas entities!
    excel_config = ExcelConfiguration()
    excel_results = from_excel(file_path, excel_config, atlas_type_defs, use_column_mapping=True)

    # Upload excel file's content to Atlas and view the guid assignments to confirm successful upload
    _upload_entities = client.upload_entities(excel_results)
    print(json.dumps(_upload_entities,indent=2))

    # Be sure to clean up the excel file stored in file_path