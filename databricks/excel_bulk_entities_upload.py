# Databricks notebook source
dbutils.fs.ls("file:/Workspace")

# COMMAND ----------

# MAGIC %pip install openpyxl pyapacheatlas

# COMMAND ----------

# DBTITLE 1,Library Imports
import pandas as pd

# COMMAND ----------

# DBTITLE 1,Get the Apache Atlas Template File Path
notebook_path = dbutils.notebook.entry_point.getDbutils().notebook().getContext().notebookPath().get()
template_folder_path = f"file:/Workspace{notebook_path.rsplit('/', 1)[0]}/source_file"
template_filename = "Metastore_paths_info.xlsx"
template_file_path = f"{template_folder_path}/{template_filename}"
print(f"template_file_path: {template_file_path}")

# COMMAND ----------

# DBTITLE 1,Read Template File into Spark Dataframe
df_template = spark.createDataFrame(pd.read_excel(template_file_path))
display(df_template)

# COMMAND ----------

import json
import os

from openpyxl import Workbook
from openpyxl import load_workbook

# PyApacheAtlas packages
# Connect to Atlas via a Service Principal
from pyapacheatlas.auth import ServicePrincipalAuthentication
from pyapacheatlas.core import PurviewClient  # Communicate with your Atlas server
from pyapacheatlas.readers import ExcelConfiguration, ExcelReader


def fill_in_workbook(filepath, excel_config):
    # You can safely ignore this function as it just
    # populates the excel spreadsheet.
    wb = load_workbook(file_path)
    bulkEntity_sheet = wb[excel_config.bulkEntity_sheet]

    # BULK Sheet SCHEMA
    #"typeName", "name", "qualifiedName"
    # Adding a couple columns to show the power of this sheet
    # [Relationship] table, type
    entities_to_load = [
        ["DataSet", "exampledataset", "pyapacheatlas://dataset",
            None, None],
        ["hive_table", "hivetable01", "pyapacheatlas://hivetable01",
            None, None],
        ["hive_column", "columnA", "pyapacheatlas://hivetable01#colA",
            'pyapacheatlas://hivetable01', 'string'],
        ["hive_column", "columnB", "pyapacheatlas://hivetable01#colB",
            'pyapacheatlas://hivetable01', 'long'],
        ["hive_column", "columnC", "pyapacheatlas://hivetable01#colC",
            'pyapacheatlas://hivetable01', 'int']
    ]

    # Need to adjust the default header to include our extra attributes
    bulkEntity_sheet['D1'] = '[Relationship] table'
    bulkEntity_sheet['E1'] = 'type'

    # Populate the excel template with samples above
    table_row_counter = 0
    for row in bulkEntity_sheet.iter_rows(min_row=2, max_col=5,
                                          max_row=len(entities_to_load) + 1):
        for idx, cell in enumerate(row):
            cell.value = entities_to_load[table_row_counter][idx]
        table_row_counter += 1

    wb.save(file_path)


if __name__ == "__main__":
    """
    This sample provides an end to end sample of reading an excel file,
    generating a batch of entities, and then uploading the entities to
    your data catalog.
    """

    # Authenticate against your Atlas server
    oauth = ServicePrincipalAuthentication(
        tenant_id=os.environ.get("TENANT_ID", ""),
        client_id=os.environ.get("CLIENT_ID", ""),
        client_secret=os.environ.get("CLIENT_SECRET", "")
    )
    client = PurviewClient(
        account_name = os.environ.get("PURVIEW_NAME", ""),
        authentication=oauth
    )

    # SETUP: This is just setting up the excel file for you
    file_path = "./demo_bulk_entities_upload.xlsx"
    excel_config = ExcelConfiguration()
    excel_reader = ExcelReader(excel_config)

    # Create an empty excel template to be populated
    excel_reader.make_template(file_path)
    # This is just a helper to fill in some demo data
    fill_in_workbook(file_path, excel_config)

    # ACTUAL WORK: This parses our excel file and creates a batch to upload
    entities = excel_reader.parse_bulk_entities(file_path)

    # This is what is getting sent to your Atlas server
    # print(json.dumps(entities,indent=2))

    results = client.upload_entities(entities)

    print(json.dumps(results, indent=2))

    print("Completed bulk upload successfully!\nSearch for hivetable01 to see your results.")

